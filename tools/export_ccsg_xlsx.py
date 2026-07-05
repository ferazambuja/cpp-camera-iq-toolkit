#!/usr/bin/env python3
"""Export the private ccsg.xlsx reference workbook to camera_iq spectral CSV.

The C++ toolkit ingests a stable text format instead of depending on an Excel
parser. This helper is the private-data bridge:

    python3 tools/export_ccsg_xlsx.py \
      data/private/references/ccsg_2019_workbook/ccsg.xlsx \
      --out data/private/references/ccsg_2019_workbook/ccsg_2_FIXED_ref.csv

Expected source sheet:
  * sheet name: ccsg_2_FIXED_ref
  * 140 rows, one per ColorChecker-SG cell
  * column B: patch ID in workbook number-major order:
    A1, B1, ... N1, A2, B2, ... N10
  * columns E..AN: reflectance bands

The workbook sheet does not carry wavelength labels. The output axis is supplied
explicitly by --first-wavelength/--step-nm/--band-count and written into the CSV
header so the C++ loader consumes an explicit axis.

The output is gitignored when written under data/private/.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


PATCH_RE = re.compile(r"^[A-N](?:[1-9]|10)$")


def expected_labels() -> list[str]:
    return [f"{col}{row}" for row in range(1, 11) for col in "ABCDEFGHIJKLMN"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", help="Private ccsg.xlsx workbook")
    parser.add_argument("--sheet", default="ccsg_2_FIXED_ref")
    parser.add_argument("--out", required=True, help="Output camera_iq spectral CSV")
    parser.add_argument("--first-wavelength", type=int, default=380)
    parser.add_argument("--step-nm", type=int, default=10)
    parser.add_argument("--band-count", type=int, default=36)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    xlsx = Path(args.xlsx)
    out = Path(args.out)
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    if args.sheet not in wb.sheetnames:
        sys.stderr.write(f"sheet not found: {args.sheet}\n")
        return 1

    ws = wb[args.sheet]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) != 140:
        sys.stderr.write(f"expected 140 rows, got {len(rows)}\n")
        return 1

    labels = [str(row[1]) for row in rows]
    if labels != expected_labels():
        sys.stderr.write(
            "patch labels are not in expected A1,B1,...N1,A2,...N10 order\n"
        )
        return 1
    if any(PATCH_RE.match(label) is None for label in labels):
        sys.stderr.write("one or more patch labels is malformed\n")
        return 1

    wavelengths = [
        args.first_wavelength + i * args.step_nm for i in range(args.band_count)
    ]
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(["patch_id", *wavelengths])
        for label, row in zip(labels, rows):
            values = row[4:4 + args.band_count]
            if len(values) != args.band_count or any(v is None for v in values):
                sys.stderr.write(f"row {label} lacks {args.band_count} bands\n")
                return 1
            writer.writerow([label, *[f"{float(v):.10g}" for v in values]])

    print(
        f"exported {len(rows)} patches x {len(wavelengths)} bands "
        f"({wavelengths[0]}-{wavelengths[-1]} nm) -> {out}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
